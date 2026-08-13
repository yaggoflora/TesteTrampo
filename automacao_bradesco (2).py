# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================

CAMINHO_EXCEL = r"C:\Users\i478606\OneDrive - Banco Bradesco S.A\Documentos\Cards-Teste.xlsx"

NOME_DA_ABA = r"Desconsiderados"

# MODO_PESQUISA define o que o script vai usar como referência:
#   "imagem" -> pesquisa apenas UMA imagem (CAMINHO_IMAGEM_PESQUISA)
#   "pasta"  -> pesquisa TODAS as imagens dentro de CAMINHO_PASTA_IMAGENS
MODO_PESQUISA = "pasta"

CAMINHO_IMAGEM_PESQUISA = r"C:\Users\i478606\Downloads\Fotos Excel\halland-com-tablet.jpg"

CAMINHO_PASTA_IMAGENS = r"C:\Users\i478606\Downloads\Fotos Excel"

EXTENSOES_VALIDAS = (".jpg", ".jpeg", ".png", ".bmp", ".webp")

SIMILARIDADE_MINIMA = 75

LINHA_INICIAL_DADOS = 2  # linha 1 = cabeçalho
COLUNA_DESCRICAO = "B"   # coluna com a descrição da imagem
COLUNA_VENCIMENTO = "F"  # coluna com a data de vencimento da peça

DIAS_DE_ALERTA = [15, 7, 3, 0]  # 0 = vence hoje


# ==============================================================================
# IMPORTAÇÕES
# ==============================================================================

import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from PIL import Image

from io import BytesIO

from datetime import datetime, date

import imagehash


# ==============================================================================
# NORMALIZA IMAGEM
# ==============================================================================

def normalizar_imagem(img):

    return img.convert("RGB")


# ==============================================================================
# HASH DA IMAGEM DE PESQUISA
# ==============================================================================

def gerar_phash_arquivo(caminho):

    with Image.open(caminho) as img:

        return imagehash.phash(
            normalizar_imagem(img)
        )


# ==============================================================================
# HASH IMAGEM DO EXCEL
# ==============================================================================

def gerar_phash_excel(img_excel):

    try:

        dados = img_excel._data()

        stream = BytesIO(dados)

        img = Image.open(stream)

        img.load()

        return imagehash.phash(
            normalizar_imagem(img)
        )

    except Exception:

        return None


# ==============================================================================
# LOCALIZAÇÃO DA IMAGEM
# ==============================================================================

def obter_celula(imagem):

    try:

        linha = imagem.anchor._from.row + 1
        coluna = imagem.anchor._from.col + 1

        return f"{get_column_letter(coluna)}{linha}"

    except Exception:

        return "Desconhecida"


# ==============================================================================
# LISTA DE IMAGENS DE REFERÊNCIA (MODO "imagem" OU "pasta")
# ==============================================================================

def obter_imagens_referencia():

    if MODO_PESQUISA == "imagem":

        return [CAMINHO_IMAGEM_PESQUISA]

    if MODO_PESQUISA == "pasta":

        if not os.path.isdir(CAMINHO_PASTA_IMAGENS):

            print(
                f"\n❌ Pasta não encontrada: {CAMINHO_PASTA_IMAGENS}"
            )

            return []

        arquivos = sorted(os.listdir(CAMINHO_PASTA_IMAGENS))

        caminhos = [
            os.path.join(CAMINHO_PASTA_IMAGENS, arquivo)
            for arquivo in arquivos
            if arquivo.lower().endswith(EXTENSOES_VALIDAS)
        ]

        return caminhos

    print(
        f"\n❌ MODO_PESQUISA inválido: '{MODO_PESQUISA}' (use 'imagem' ou 'pasta')"
    )

    return []


# ==============================================================================
# STATUS DE VENCIMENTO
# ==============================================================================

def classificar_vencimento(data_vencimento):

    if not isinstance(data_vencimento, (datetime, date)):
        return None

    if isinstance(data_vencimento, datetime):
        data_vencimento = data_vencimento.date()

    hoje = date.today()

    dias_restantes = (data_vencimento - hoje).days

    if dias_restantes < 0:
        return "VENCIDA"

    if dias_restantes == 0:
        return "VENCE HOJE"

    if dias_restantes in DIAS_DE_ALERTA:
        return f"VENCE EM {dias_restantes} DIA(S)"

    return None


# ==============================================================================
# EXTRAI O NÚMERO DA LINHA A PARTIR DE UMA CÉLULA (EX: "C15" -> 15)
# ==============================================================================

def extrair_linha(celula):

    numero = ""

    for caractere in celula:

        if caractere.isdigit():
            numero += caractere

    return int(numero) if numero else None


# ==============================================================================
# PESQUISA DE UMA ÚNICA IMAGEM DE REFERÊNCIA DENTRO DA ABA
# ==============================================================================

def pesquisar_uma_imagem(hash_pesquisa, aba):

    encontrados_identicos = []

    encontrados_semelhantes = []

    for imagem in aba._images:

        hash_excel = gerar_phash_excel(imagem)

        if hash_excel is None:
            continue

        distancia = hash_pesquisa - hash_excel

        similaridade = (
            (64 - distancia) / 64
        ) * 100

        celula = obter_celula(imagem)

        if distancia == 0:
            encontrados_identicos.append({
                "celula": celula,
                "distancia": distancia
            })
        elif similaridade >= SIMILARIDADE_MINIMA:
            encontrados_semelhantes.append({
                "celula": celula,
                "distancia": distancia,
                "similaridade": round(similaridade, 2)
            })

    if encontrados_identicos:
        return encontrados_identicos

    return encontrados_semelhantes


# ==============================================================================
# MONTA O RESULTADO COMPLETO DE UMA LINHA (DESCRIÇÃO + VENCIMENTO)
# ==============================================================================

def montar_info_linha(aba, linha):

    indice_descricao = column_index_from_string(COLUNA_DESCRICAO)
    indice_vencimento = column_index_from_string(COLUNA_VENCIMENTO)

    descricao = aba.cell(row=linha, column=indice_descricao).value

    data_vencimento = aba.cell(row=linha, column=indice_vencimento).value

    status_vencimento = classificar_vencimento(data_vencimento)

    data_formatada = (
        data_vencimento.strftime("%d/%m/%Y")
        if isinstance(data_vencimento, (datetime, date))
        else data_vencimento
    )

    return {
        "descricao": descricao if descricao else "(sem descrição)",
        "data_vencimento": data_formatada,
        "status_vencimento": status_vencimento
    }


# ==============================================================================
# EXECUÇÃO PRINCIPAL - PESQUISA + DESCRIÇÃO + VENCIMENTO
# ==============================================================================

def pesquisar_imagens():

    print("\nAbrindo planilha...")

    wb = load_workbook(
        CAMINHO_EXCEL,
        data_only=True
    )

    if NOME_DA_ABA not in wb.sheetnames:

        print(
            f"\n❌ Aba '{NOME_DA_ABA}' não encontrada."
        )

        return

    aba = wb[NOME_DA_ABA]

    if not aba._images:

        print(
            "\n❌ Nenhuma imagem encontrada na aba."
        )

        return

    imagens_referencia = obter_imagens_referencia()

    if not imagens_referencia:

        print(
            "\n❌ Nenhuma imagem de referência para pesquisar."
        )

        return

    print(
        f"\nImagens de referência a pesquisar: {len(imagens_referencia)}"
    )

    print(
        f"Imagens encontradas na planilha: {len(aba._images)}"
    )

    print("\n" + "="*50)
    print("📋 RESULTADO")
    print("="*50)

    for caminho_imagem in imagens_referencia:

        nome_arquivo = os.path.basename(caminho_imagem)

        print(f"\n🔎 Referência: {nome_arquivo}")

        try:
            hash_pesquisa = gerar_phash_arquivo(caminho_imagem)
        except Exception:
            print("   ❌ Não foi possível abrir a imagem de referência.")
            continue

        encontrados = pesquisar_uma_imagem(hash_pesquisa, aba)

        if not encontrados:
            print("   ❌ Não encontrada na planilha.")
            continue

        for item in encontrados:

            linha = extrair_linha(item["celula"])

            if linha is None:
                continue

            info = montar_info_linha(aba, linha)

            similaridade_txt = (
                f" | Similaridade: {item['similaridade']}%"
                if "similaridade" in item
                else " | Idêntica"
            )

            print(
                f"   📍 Célula: {item['celula']}{similaridade_txt}"
            )
            print(
                f"      📝 Descrição: {info['descricao']}"
            )

            if info["status_vencimento"]:
                print(
                    f"      ⚠️ Vencimento: {info['data_vencimento']} | Status: {info['status_vencimento']}"
                )
            else:
                print(
                    f"      ✅ Vencimento: {info['data_vencimento']} | Sem alerta de vencimento"
                )

    print("\n" + "="*50)


# ==============================================================================
# EXECUÇÃO
# ==============================================================================

if __name__ == "__main__":

    pesquisar_imagens()
