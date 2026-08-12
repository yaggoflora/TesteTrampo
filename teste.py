"""
=================================================================================================
 SCRIPT: comparar_imagens_excel.py
 OBJETIVO:
     Comparar uma imagem de referencia (arquivo local) com as imagens flutuantes presentes em
     uma coluna especifica de uma aba de um arquivo Excel (.xlsx), classificando cada uma como:
         - "Identica"          -> bytes exatamente iguais (hash SHA256 identico)
         - "Muito semelhante"  -> visualmente quase igual, mas nao byte-a-byte igual
         - "Semelhante"        -> visualmente parecida, acima do limiar minimo
         - (nada)               -> abaixo do limiar minimo -> NAO aparece no resultado

 Equivalente em Python do script PowerShell "Comparar-Imagens-Excel.ps1 (v3)", usando:
     - openpyxl  -> le o .xlsx e extrai as imagens flutuantes com seus bytes ORIGINAIS
                    (sem re-renderizar) e a celula de ancoragem, exatamente como a v3 fazia
                    lendo o XML do ZIP manualmente.
     - Pillow    -> abre as imagens e calcula a miniatura reduzida usada na similaridade visual.
     - pandas    -> monta a tabela final de resultados e permite exportar para .xlsx/.csv.

 REGRA DE DECISAO (igual ao script original), nesta ordem:
   1) Se o hash SHA256 dos bytes originais bater EXATAMENTE  -> Status = "Identica" (100%)
   2) Senao, calcula um percentual de similaridade visual (0 a 100) comparando uma grade
      reduzida de pixels das duas imagens:
        percentual >= LIMIAR_MUITO_SEMELHANTE  -> Status = "Muito semelhante"
        percentual >= LIMIAR_SEMELHANTE        -> Status = "Semelhante"
        percentual <  LIMIAR_SEMELHANTE        -> imagem e DESCARTADA (nao entra no resultado)
=================================================================================================
"""

from __future__ import annotations

import hashlib
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, UnidentifiedImageError

# =================================================================================================
# CONFIGURACOES DO USUARIO -- AJUSTE SOMENTE ESTA SECAO
# =================================================================================================

# Caminho completo do arquivo Excel a ser analisado
CAMINHO_ARQUIVO_EXCEL = r"C:\Caminho\Para\Arquivo.xlsx"

# Nome exato da aba (planilha) onde estao as imagens
NOME_ABA = "Planilha1"

# Caminho completo da imagem de referencia que sera usada na comparacao
CAMINHO_IMAGEM_REFERENCIA = r"C:\Caminho\Para\ImagemReferencia.png"

# Numero da coluna que deve ser considerada (Coluna B = 2, Coluna A = 1, Coluna C = 3, etc)
COLUNA_ALVO = 2

# -------------------------------------------------------------------------------------------
# >>> LIMIARES DE CLASSIFICACAO DE SEMELHANCA (edite aqui para ajustar o comportamento) <<<
# -------------------------------------------------------------------------------------------
LIMIAR_MUITO_SEMELHANTE = 95   # percentual minimo (%) para ser "Muito semelhante"
LIMIAR_SEMELHANTE = 80         # percentual minimo (%) para ser "Semelhante"

# Resolucao da grade usada no calculo de similaridade visual (largura x altura, em "pixels
# logicos" reduzidos). Valores maiores = comparacao mais precisa, porem mais lenta.
# 16 e um bom equilibrio entre precisao e performance para 1000+ imagens.
RESOLUCAO_GRADE_COMPARACAO = 16

# Caminho para exportar o resultado final (None = nao exporta, so mostra no console)
CAMINHO_SAIDA_XLSX: Optional[str] = None  # ex: r"C:\Caminho\Para\resultado.xlsx"

# =================================================================================================
# NAO E NECESSARIO ALTERAR NADA A PARTIR DAQUI
# =================================================================================================


@dataclass
class ResultadoComparacao:
    celula: str
    status: str
    similaridade: float
    metodo: str = "Shape flutuante"


def get_hash_de_bytes(dados: bytes) -> str:
    """SHA256 dos bytes crus -- equivalente a Get-HashDeBytes."""
    return hashlib.sha256(dados).hexdigest().upper()


def converter_para_miniatura_comparacao(imagem: Image.Image, tamanho: int) -> Image.Image:
    """
    Reduz uma imagem a uma grade NxN (forca redimensionamento, ignora proporcao original)
    -- equivalente a ConvertTo-MiniaturaComparacao. Usa bicubic para ficar igual ao
    HighQualityBicubic do GDI+.
    """
    return imagem.convert("RGB").resize((tamanho, tamanho), Image.BICUBIC)


def get_percentual_similaridade(imagem1: Image.Image, imagem2: Image.Image, tamanho: int) -> float:
    """
    Calcula o percentual de similaridade visual (0 a 100) entre duas imagens, reduzindo
    ambas a uma grade pequena de pixels e comparando a diferenca media de cor (R,G,B)
    -- equivalente a Get-PercentualSimilaridade.
    """
    miniatura1 = converter_para_miniatura_comparacao(imagem1, tamanho)
    miniatura2 = converter_para_miniatura_comparacao(imagem2, tamanho)

    pixels1 = miniatura1.load()
    pixels2 = miniatura2.load()

    soma_diferencas = 0.0
    for x in range(tamanho):
        for y in range(tamanho):
            r1, g1, b1 = pixels1[x, y]
            r2, g2, b2 = pixels2[x, y]
            soma_diferencas += (abs(r1 - r2) + abs(g1 - g2) + abs(b1 - b2)) / 3.0

    total_pontos = tamanho * tamanho
    diferenca_media = soma_diferencas / total_pontos  # 0 (identico) a 255 (totalmente oposto)
    percentual = 100 - ((diferenca_media / 255) * 100)
    return round(percentual, 2)


def get_resultado_comparacao(
    bytes_candidato: bytes,
    celula: str,
    bytes_referencia: bytes,
    hash_referencia: str,
    imagem_referencia_carregada: Image.Image,
) -> Optional[ResultadoComparacao]:
    """
    Processa um candidato (bytes de uma imagem encontrada na planilha) e devolve o resultado
    classificado, ou None se ficar abaixo do limiar minimo de semelhanca
    -- equivalente a Get-ResultadoComparacao.
    """
    hash_candidato = get_hash_de_bytes(bytes_candidato)

    if hash_candidato == hash_referencia:
        return ResultadoComparacao(celula=celula, status="Identica", similaridade=100.0)

    try:
        imagem_candidata = Image.open(pd.io.common.BytesIO(bytes_candidato))
        imagem_candidata.load()
    except (UnidentifiedImageError, OSError):
        # Arquivo nao e uma imagem valida / formato nao suportado pelo Pillow
        return None

    percentual = get_percentual_similaridade(
        imagem_referencia_carregada, imagem_candidata, RESOLUCAO_GRADE_COMPARACAO
    )

    if percentual >= LIMIAR_MUITO_SEMELHANTE:
        status = "Muito semelhante"
    elif percentual >= LIMIAR_SEMELHANTE:
        status = "Semelhante"
    else:
        return None  # Abaixo do limiar minimo -> nao retorna nada (regra de negocio)

    return ResultadoComparacao(celula=celula, status=status, similaridade=percentual)


def extrair_imagens_da_coluna(caminho_excel: str, nome_aba: str, coluna_alvo: int):
    """
    Abre o .xlsx com openpyxl e devolve uma lista de tuplas (endereco_celula, bytes_imagem)
    para cada imagem flutuante ancorada na coluna alvo da aba informada.

    Assim como a v3 do script PowerShell le o XML do desenho (drawingN.xml) para pegar os
    bytes originais sem re-renderizar, o openpyxl tambem le os bytes originais do arquivo
    dentro de xl/media ao carregar a planilha (nao ha nenhuma renderizacao via COM/Clipboard).
    """
    wb = load_workbook(caminho_excel, data_only=True)
    if nome_aba not in wb.sheetnames:
        raise ValueError(f"A aba '{nome_aba}' nao foi encontrada no arquivo. Confira o nome exato.")

    ws = wb[nome_aba]
    imagens_encontradas = getattr(ws, "_images", [])

    candidatos = []
    for img in imagens_encontradas:
        ancora = img.anchor
        # OneCellAnchor e TwoCellAnchor ambos possuem o atributo _from com col/row (base zero)
        marcador_from = getattr(ancora, "_from", None)
        if marcador_from is None:
            continue

        coluna_real = marcador_from.col + 1  # base zero -> base um
        linha_real = marcador_from.row + 1

        if coluna_real != coluna_alvo:
            continue

        endereco_celula = f"{get_column_letter(coluna_real)}{linha_real}"
        candidatos.append((endereco_celula, img._data()))

    wb.close()
    return candidatos


def comparar_imagens_excel() -> pd.DataFrame:
    print("=" * 70)
    print(" Iniciando comparacao de imagens - Excel x Imagem de Referencia")
    print("=" * 70)

    caminho_excel = Path(CAMINHO_ARQUIVO_EXCEL)
    caminho_referencia = Path(CAMINHO_IMAGEM_REFERENCIA)

    if not caminho_excel.is_file():
        sys.exit(f"Arquivo Excel nao encontrado em: {caminho_excel}")
    if not caminho_referencia.is_file():
        sys.exit(f"Imagem de referencia nao encontrada em: {caminho_referencia}")

    # ------------------------------------------------------------------------------------
    # ETAPA 1: Carrega a imagem de referencia (uma unica vez)
    # ------------------------------------------------------------------------------------
    print("\n[1/4] Carregando imagem de referencia...")

    bytes_referencia = caminho_referencia.read_bytes()
    hash_referencia = get_hash_de_bytes(bytes_referencia)
    imagem_referencia_carregada = Image.open(pd.io.common.BytesIO(bytes_referencia))
    imagem_referencia_carregada.load()

    print(f"      Hash SHA256 : {hash_referencia}")
    print(f"      Dimensoes   : {imagem_referencia_carregada.width}x{imagem_referencia_carregada.height}")

    # ------------------------------------------------------------------------------------
    # ETAPA 2: Abre o .xlsx e localiza as imagens flutuantes da coluna alvo
    # ------------------------------------------------------------------------------------
    print(f"\n[2/4] Localizando a aba '{NOME_ABA}' dentro do arquivo...")

    candidatos = extrair_imagens_da_coluna(str(caminho_excel), NOME_ABA, COLUNA_ALVO)
    print(f"      Imagens flutuantes encontradas na coluna alvo: {len(candidatos)}")

    # ------------------------------------------------------------------------------------
    # ETAPA 3: Classifica cada candidato
    # ------------------------------------------------------------------------------------
    print("\n[3/4] Comparando imagens...")

    resultados: list[ResultadoComparacao] = []
    total = len(candidatos)
    for indice, (celula, bytes_candidato) in enumerate(candidatos, start=1):
        if indice % 20 == 0 or indice == total:
            print(f"      {indice}/{total} ({round(indice / total * 100, 1)}%)")

        resultado = get_resultado_comparacao(
            bytes_candidato=bytes_candidato,
            celula=celula,
            bytes_referencia=bytes_referencia,
            hash_referencia=hash_referencia,
            imagem_referencia_carregada=imagem_referencia_carregada,
        )
        if resultado is not None:
            resultados.append(resultado)

    # ------------------------------------------------------------------------------------
    # ETAPA 4: Monta e exibe a tabela final
    # ------------------------------------------------------------------------------------
    print("\n[4/4] Resultado final:")
    print("\n" + "=" * 70)
    print(" RESULTADO DA COMPARACAO")
    print("=" * 70)

    df = pd.DataFrame(
        [
            {
                "Celula": r.celula,
                "Status": r.status,
                "Similaridade": r.similaridade,
                "Metodo": r.metodo,
            }
            for r in resultados
        ]
    )

    if df.empty:
        print(f"Nenhuma imagem identica ou semelhante (>= {LIMIAR_SEMELHANTE}%) foi encontrada.")
    else:
        df = df.sort_values("Celula").reset_index(drop=True)
        print(df.to_string(index=False))

        if CAMINHO_SAIDA_XLSX:
            df.to_excel(CAMINHO_SAIDA_XLSX, index=False)
            print(f"\nResultado exportado para: {CAMINHO_SAIDA_XLSX}")

    print(f"\nProcesso concluido. Total de correspondencias encontradas: {len(df)}")
    return df


if __name__ == "__main__":
    comparar_imagens_excel()