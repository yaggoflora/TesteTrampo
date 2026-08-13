# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================

CAMINHO_EXCEL = r"C:\Users\i478606\OneDrive - Banco Bradesco S.A\Documentos\Cards-Teste.xlsx"

NOME_DA_ABA = r"Desconsiderados"

CAMINHO_IMAGEM_PESQUISA = r"C:\Users\i478606\Downloads\Fotos Excel\halland-com-tablet.jpg"
SIMILARIDADE_MINIMA = 75

LINHA_INICIAL_DADOS = 2  # linha 1 = cabeçalho
COLUNA_VENCIMENTO = "F"  # coluna com a data de vencimento da peça

DIAS_DE_ALERTA = [15, 7, 3, 0]  # 0 = vence hoje


# ==============================================================================
# IMPORTAÇÕES
# ==============================================================================

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from PIL import Image

from io import BytesIO

from datetime import datetime, date

import imagehash


# ==============================================================================
# NORMALIZA IMAGEM
# ==============================================================================

def normalizar_imagem(img):

    return img.convert("RGB")


# ==============================================================================
# HASH DA IMAGEM DE PESQUISA
# ==============================================================================

def gerar_phash_arquivo(caminho):

    with Image.open(caminho) as img:

        return imagehash.phash(
            normalizar_imagem(img)
        )


# ==============================================================================
# HASH IMAGEM DO EXCEL
# ==============================================================================

def gerar_phash_excel(img_excel):

    try:

        dados = img_excel._data()

        stream = BytesIO(dados)

        img = Image.open(stream)

        img.load()

        return imagehash.phash(
            normalizar_imagem(img)
        )

    except Exception:

        return None


# ==============================================================================
# LOCALIZAÇÃO DA IMAGEM
# ==============================================================================

def obter_celula(imagem):

    try:

        linha = imagem.anchor._from.row + 1
        coluna = imagem.anchor._from.col + 1

        return f"{get_column_letter(coluna)}{linha}"

    except Exception:

        return "Desconhecida"


# ==============================================================================
# PESQUISA
# ==============================================================================

def pesquisar_imagem():

    print("\nAbrindo planilha...")

    wb = load_workbook(
        CAMINHO_EXCEL,
        data_only=True
    )

    if NOME_DA_ABA not in wb.sheetnames:

        print(
            f"\n❌ Aba '{NOME_DA_ABA}' não encontrada."
        )

        return

    aba = wb[NOME_DA_ABA]

    if not aba._images:

        print(
            "\n❌ Nenhuma imagem encontrada na aba."
        )

        return

    hash_pesquisa = gerar_phash_arquivo(
        CAMINHO_IMAGEM_PESQUISA
    )

    encontrados_identicos = []

    encontrados_semelhantes = []

    print(
        f"\nImagens encontradas: {len(aba._images)}"
    )

    for imagem in aba._images:

        hash_excel = gerar_phash_excel(imagem)

        if hash_excel is None:
            continue

        distancia = hash_pesquisa - hash_excel

        similaridade = (
            (64 - distancia) / 64
        ) * 100

        celula = obter_celula(imagem)

        if distancia == 0:
            encontrados_identicos.append({
                "celula": celula,
                "distancia": distancia
            })
        elif similaridade >= SIMILARIDADE_MINIMA:
            encontrados_semelhantes.append({
                "celula": celula,
                "distancia": distancia,
                "similaridade": round(similaridade, 2)
            })

    print("\n" + "="*50)
    print("📋 RESULTADO")
    print("="*50)

    if encontrados_identicos:

        print(
            "\n✅ IMAGEM PRESENTE (IDÊNTICA)"
        )

        for item in encontrados_identicos:

            print(
                f"📍 Célula: {item['celula']}"
            )

        return

    if encontrados_semelhantes:

        print(
            "\n⚠️ IMAGEM SEMELHANTE ENCONTRADA"
        )

        for item in encontrados_semelhantes:

            print(
                f"📍 Célula: {item['celula']} | Similaridade: {item['similaridade']}%"
            )

        return

    print(
        "\n❌ IMAGEM NÃO CONTÉM NA PLANILHA"
    )

    print("="*50)


# ==============================================================================
# STATUS DE VENCIMENTO
# ==============================================================================

def classificar_vencimento(data_vencimento):

    if not isinstance(data_vencimento, (datetime, date)):
        return None

    if isinstance(data_vencimento, datetime):
        data_vencimento = data_vencimento.date()

    hoje = date.today()

    dias_restantes = (data_vencimento - hoje).days

    if dias_restantes < 0:
        return "VENCIDA"

    if dias_restantes == 0:
        return "VENCE HOJE"

    if dias_restantes in DIAS_DE_ALERTA:
        return f"VENCE EM {dias_restantes} DIA(S)"

    return None


# ==============================================================================
# VERIFICAÇÃO DE PEÇAS VENCIDAS / PRESTES A VENCER
# ==============================================================================

def verificar_vencimentos():

    print("\nVerificando vencimentos...")

    wb = load_workbook(
        CAMINHO_EXCEL,
        data_only=True
    )

    if NOME_DA_ABA not in wb.sheetnames:

        print(
            f"\n❌ Aba '{NOME_DA_ABA}' não encontrada."
        )

        return

    aba = wb[NOME_DA_ABA]

    indice_coluna = column_index_from_string(COLUNA_VENCIMENTO)

    encontrados = []

    for linha in range(LINHA_INICIAL_DADOS, aba.max_row + 1):

        celula = aba.cell(row=linha, column=indice_coluna)

        status = classificar_vencimento(celula.value)

        if status:

            encontrados.append({
                "celula": f"{COLUNA_VENCIMENTO}{linha}",
                "data": celula.value,
                "status": status
            })

    print("\n" + "="*50)
    print("📅 RESULTADO - VENCIMENTOS")
    print("="*50)

    if not encontrados:

        print(
            "\n✅ Nenhuma peça vencida ou prestes a vencer."
        )

        print("="*50)

        return

    for item in encontrados:

        data_formatada = item["data"].strftime("%d/%m/%Y") if isinstance(item["data"], (datetime, date)) else item["data"]

        print(
            f"⚠️ Célula: {item['celula']} | Data: {data_formatada} | Status: {item['status']}"
        )

    print("="*50)


# ==============================================================================
# EXECUÇÃO
# ==============================================================================

if __name__ == "__main__":

    pesquisar_imagem()

    verificar_vencimentos()
