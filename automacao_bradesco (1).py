# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================

CAMINHO_EXCEL = r"C:\Users\i478606\OneDrive - Banco Bradesco S.A\Documentos\Cards-Teste.xlsx"

NOME_DA_ABA = r"Desconsiderados"

CAMINHO_IMAGEM_PESQUISA = r"C:\Users\i478606\Downloads\Fotos Excel\halland-com-tablet.jpg"
SIMILARIDADE_MINIMA = 75

LINHA_INICIAL_DADOS = 2  # linha 1 = cabeçalho
COLUNA_VENCIMENTO = "F"  # coluna com a data de vencimento da peça

DIAS_DE_ALERTA = [15, 7, 3, 0]  # 0 = vence hoje


# ==============================================================================
# IMPORTAÇÕES
# ==============================================================================

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from PIL import Image

from io import BytesIO

from datetime import datetime, date

import imagehash


# ==============================================================================
# NORMALIZA IMAGEM
# ==============================================================================

def normalizar_imagem(img):

    return img.convert("RGB")


# ==============================================================================
# HASH DA IMAGEM DE PESQUISA
# ==============================================================================

def gerar_phash_arquivo(caminho):

    with Image.open(caminho) as img:

        return imagehash.phash(
            normalizar_imagem(img)
        )


# ==============================================================================
# HASH IMAGEM DO EXCEL
# ==============================================================================

def gerar_phash_excel(img_excel):

    try:

        dados = img_excel._data()

        stream = BytesIO(dados)

        img = Image.open(stream)

        img.load()

        return imagehash.phash(
            normalizar_imagem(img)
        )

    except Exception:

        return None


# ==============================================================================
# LOCALIZAÇÃO DA IMAGEM
# ==============================================================================

def obter_celula(imagem):

    try:

        linha = imagem.anchor._from.row + 1
        coluna = imagem.anchor._from.col + 1

        return f"{get_column_letter(coluna)}{linha}"

    except Exception:

        return "Desconhecida"


# ==============================================================================
# PESQUISA
# ==============================================================================

def pesquisar_imagem():

    print("\nAbrindo planilha...")

    wb = load_workbook(
        CAMINHO_EXCEL,
        data_only=True
    )

    if NOME_DA_ABA not in wb.sheetnames:

        print(
            f"\n❌ Aba '{NOME_DA_ABA}' não encontrada."
        )

        return []

    aba = wb[NOME_DA_ABA]

    if not aba._images:

        print(
            "\n❌ Nenhuma imagem encontrada na aba."
        )

        return []

    hash_pesquisa = gerar_phash_arquivo(
        CAMINHO_IMAGEM_PESQUISA
    )

    encontrados_identicos = []

    encontrados_semelhantes = []

    print(
        f"\nImagens encontradas: {len(aba._images)}"
    )

    for imagem in aba._images:

        hash_excel = gerar_phash_excel(imagem)

        if hash_excel is None:
            continue

        distancia = hash_pesquisa - hash_excel

        similaridade = (
            (64 - distancia) / 64
        ) * 100

        celula = obter_celula(imagem)

        if distancia == 0:
            encontrados_identicos.append({
                "celula": celula,
                "distancia": distancia
            })
        elif similaridade >= SIMILARIDADE_MINIMA:
            encontrados_semelhantes.append({
                "celula": celula,
                "distancia": distancia,
                "similaridade": round(similaridade, 2)
            })

    print("\n" + "="*50)
    print("📋 RESULTADO")
    print("="*50)

    if encontrados_identicos:

        print(
            "\n✅ IMAGEM PRESENTE (IDÊNTICA)"
        )

        for item in encontrados_identicos:

            print(
                f"📍 Célula: {item['celula']}"
            )

        print("="*50)

        return encontrados_identicos

    if encontrados_semelhantes:

        print(
            "\n⚠️ IMAGEM SEMELHANTE ENCONTRADA"
        )

        for item in encontrados_semelhantes:

            print(
                f"📍 Célula: {item['celula']} | Similaridade: {item['similaridade']}%"
            )

        print("="*50)

        return encontrados_semelhantes

    print(
        "\n❌ IMAGEM NÃO CONTÉM NA PLANILHA"
    )

    print("="*50)

    return []


# ==============================================================================
# STATUS DE VENCIMENTO
# ==============================================================================

def classificar_vencimento(data_vencimento):

    if not isinstance(data_vencimento, (datetime, date)):
        return None

    if isinstance(data_vencimento, datetime):
        data_vencimento = data_vencimento.date()

    hoje = date.today()

    dias_restantes = (data_vencimento - hoje).days

    if dias_restantes < 0:
        return "VENCIDA"

    if dias_restantes == 0:
        return "VENCE HOJE"

    if dias_restantes in DIAS_DE_ALERTA:
        return f"VENCE EM {dias_restantes} DIA(S)"

    return None


# ==============================================================================
# EXTRAI O NÚMERO DA LINHA A PARTIR DE UMA CÉLULA (EX: "C15" -> 15)
# ==============================================================================

def extrair_linha(celula):

    numero = ""

    for caractere in celula:

        if caractere.isdigit():
            numero += caractere

    return int(numero) if numero else None


# ==============================================================================
# VERIFICAÇÃO DE PEÇAS VENCIDAS / PRESTES A VENCER
# (somente nas linhas onde a imagem pesquisada foi encontrada)
# ==============================================================================

def verificar_vencimentos(itens_encontrados):

    if not itens_encontrados:

        print(
            "\nNenhuma ocorrência da imagem pesquisada para checar vencimento."
        )

        return

    print("\nVerificando vencimento da(s) peça(s) encontrada(s)...")

    wb = load_workbook(
        CAMINHO_EXCEL,
        data_only=True
    )

    if NOME_DA_ABA not in wb.sheetnames:

        print(
            f"\n❌ Aba '{NOME_DA_ABA}' não encontrada."
        )

        return

    aba = wb[NOME_DA_ABA]

    indice_coluna = column_index_from_string(COLUNA_VENCIMENTO)

    print("\n" + "="*50)
    print("📅 RESULTADO - VENCIMENTO")
    print("="*50)

    for item in itens_encontrados:

        linha = extrair_linha(item["celula"])

        if linha is None:
            continue

        celula_vencimento = aba.cell(row=linha, column=indice_coluna)

        data_vencimento = celula_vencimento.value

        status = classificar_vencimento(data_vencimento)

        data_formatada = data_vencimento.strftime("%d/%m/%Y") if isinstance(data_vencimento, (datetime, date)) else data_vencimento

        if status:
            print(
                f"⚠️ Imagem em {item['celula']} | Vencimento ({COLUNA_VENCIMENTO}{linha}): {data_formatada} | Status: {status}"
            )
        else:
            print(
                f"✅ Imagem em {item['celula']} | Vencimento ({COLUNA_VENCIMENTO}{linha}): {data_formatada} | Sem alerta de vencimento"
            )

    print("="*50)


# ==============================================================================
# EXECUÇÃO
# ==============================================================================

if __name__ == "__main__":

    itens_encontrados = pesquisar_imagem()

    verificar_vencimentos(itens_encontrados)
