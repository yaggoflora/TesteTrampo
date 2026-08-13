# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================

CAMINHO_EXCEL = r"C:\Users\i478606\OneDrive - Banco Bradesco S.A\Documentos\Cards-Teste.xlsx"

NOME_DA_ABA = r"Desconsiderados"

CAMINHO_IMAGEM_PESQUISA = r"C:\Users\i478606\Downloads\Fotos Excel\halland-com-tablet.jpg"
SIMILARIDADE_MINIMA = 75

LINHA_INICIAL_DADOS = 2  # linha 1 = cabeçalho
COLUNA_DESCRICAO = "B"   # coluna com a descrição da peça
COLUNA_VENCIMENTO = "F"  # coluna com a data de vencimento da peça

DIAS_DE_ALERTA = [15, 7, 3, 0]  # 0 = vence hoje


# ==============================================================================
# IMPORTAÇÕES
# ==============================================================================

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from PIL import Image
from io import BytesIO
from datetime import datetime, date

import imagehash


# ==============================================================================
# NORMALIZA IMAGEM
# ==============================================================================

def normalizar_imagem(img):
    return img.convert("RGB")


# ==============================================================================
# HASH DA IMAGEM DE PESQUISA
# ==============================================================================

def gerar_phash_arquivo(caminho):
    with Image.open(caminho) as img:
        return imagehash.phash(normalizar_imagem(img))


# ==============================================================================
# HASH IMAGEM DO EXCEL
# ==============================================================================

def gerar_phash_excel(img_excel):
    """
    Retorna o phash da imagem embutida no Excel.
    Se a imagem estiver corrompida/em formato não suportado, a exceção
    sobe para quem chamou decidir o que fazer (sem derrubar o script todo).
    """
    dados = img_excel._data()
    stream = BytesIO(dados)
    img = Image.open(stream)
    img.load()
    return imagehash.phash(normalizar_imagem(img))


# ==============================================================================
# LOCALIZAÇÃO DA IMAGEM (linha e coluna) — robusto a diferentes tipos de âncora
# ==============================================================================

def obter_linha_coluna(imagem):
    """
    openpyxl pode representar a âncora da imagem como OneCellAnchor ou
    TwoCellAnchor. Em ambos os casos existe o atributo `_from`, mas
    tratamos com segurança para não quebrar o script se algum dia vier
    outro formato de âncora.
    """
    anchor = imagem.anchor

    frm = getattr(anchor, "_from", None)
    if frm is None:
        raise ValueError("Âncora da imagem não possui referência de célula (_from).")

    linha = frm.row + 1
    coluna = frm.col + 1

    return linha, coluna


# ==============================================================================
# LEITURA DE CÉLULA CONSIDERANDO CÉLULAS MESCLADAS
# ==============================================================================

def obter_valor_real(aba, linha, coluna_letra):
    """
    Se a célula pedida faz parte de um intervalo mesclado, o Excel só guarda
    o valor na célula "mestre" (canto superior esquerdo do range). Esta
    função resolve isso automaticamente, buscando o valor correto mesmo que
    a célula informada não seja a mestre do merge.
    """
    endereco = f"{coluna_letra}{linha}"
    celula = aba[endereco]

    for intervalo in aba.merged_cells.ranges:
        if celula.coordinate in intervalo:
            return aba.cell(row=intervalo.min_row, column=intervalo.min_col).value

    return celula.value


# ==============================================================================
# STATUS DE VENCIMENTO
# ==============================================================================

def classificar_vencimento(data_vencimento):

    if not isinstance(data_vencimento, (datetime, date)):
        return "SEM DATA / OK"

    if isinstance(data_vencimento, datetime):
        data_vencimento = data_vencimento.date()

    hoje = date.today()
    dias_restantes = (data_vencimento - hoje).days

    if dias_restantes < 0:
        return "VENCIDA"

    if dias_restantes == 0:
        return "VENCE HOJE"

    if dias_restantes in DIAS_DE_ALERTA:
        return f"VENCE EM {dias_restantes} DIA(S)"

    return "OK (fora da janela de alerta)"


# ==============================================================================
# PESQUISA — retorna TODAS as ocorrências da imagem pesquisada
# ==============================================================================

def pesquisar_imagem():

    print("\nAbrindo planilha...")

    wb = load_workbook(CAMINHO_EXCEL, data_only=True)

    if NOME_DA_ABA not in wb.sheetnames:
        print(f"\n❌ Aba '{NOME_DA_ABA}' não encontrada.")
        return

    aba = wb[NOME_DA_ABA]

    if not aba._images:
        print("\n❌ Nenhuma imagem encontrada na aba.")
        return

    hash_pesquisa = gerar_phash_arquivo(CAMINHO_IMAGEM_PESQUISA)

    total_imagens = len(aba._images)
    print(f"\nImagens encontradas na planilha: {total_imagens}")
    print("Comparando com a imagem de pesquisa...\n")

    col_desc = column_index_from_string(COLUNA_DESCRICAO)
    col_venc = column_index_from_string(COLUNA_VENCIMENTO)

    encontrados_identicos = []
    encontrados_semelhantes = []
    falhas = []

    for indice, imagem in enumerate(aba._images, start=1):

        # cada imagem é processada isoladamente: se uma falhar, o script
        # avisa e segue para a próxima, em vez de travar tudo
        try:
            linha, coluna = obter_linha_coluna(imagem)
        except Exception as erro:
            falhas.append(f"Imagem #{indice}: não foi possível localizar a célula ({erro})")
            continue

        try:
            hash_excel = gerar_phash_excel(imagem)
        except Exception as erro:
            falhas.append(f"Imagem #{indice} (linha {linha}): falha ao ler/decodificar a imagem ({erro})")
            continue

        distancia = hash_pesquisa - hash_excel
        similaridade = round(((64 - distancia) / 64) * 100, 2)

        # não é match suficiente -> ignora e segue
        if distancia != 0 and similaridade < SIMILARIDADE_MINIMA:
            continue

        # daqui pra baixo é um match (idêntico ou semelhante) -> busca
        # descrição e vencimento SÓ dessa linha, já resolvendo merge
        descricao = obter_valor_real(aba, linha, COLUNA_DESCRICAO)
        data_vencimento = obter_valor_real(aba, linha, COLUNA_VENCIMENTO)
        status_vencimento = classificar_vencimento(data_vencimento)

        data_formatada = (
            data_vencimento.strftime("%d/%m/%Y")
            if isinstance(data_vencimento, (datetime, date))
            else data_vencimento
        )

        registro = {
            "celula": f"{get_column_letter(coluna)}{linha}",
            "linha": linha,
            "distancia": distancia,
            "similaridade": similaridade,
            "descricao": descricao,
            "vencimento": data_formatada,
            "status_vencimento": status_vencimento,
        }

        if distancia == 0:
            encontrados_identicos.append(registro)
        else:
            encontrados_semelhantes.append(registro)

    # ==========================================================================
    # RELATÓRIO FINAL
    # ==========================================================================

    print("="*60)
    print("📋 RESULTADO DA BUSCA")
    print("="*60)

    total_matches = len(encontrados_identicos) + len(encontrados_semelhantes)

    if total_matches == 0:
        print("\n❌ Nenhuma ocorrência da imagem foi encontrada na planilha.")

    else:
        print(f"\n✅ {total_matches} ocorrência(s) encontrada(s):\n")

        # ordena por linha, pra facilitar conferência visual na planilha
        todos = sorted(encontrados_identicos + encontrados_semelhantes, key=lambda x: x["linha"])

        for item in todos:
            tag = "✅ IDÊNTICA" if item["distancia"] == 0 else f"⚠️ SEMELHANTE ({item['similaridade']}%)"
            print(f"{tag}")
            print(f"   📍 Célula da imagem : {item['celula']}")
            print(f"   📝 Descrição        : {item['descricao']}")
            print(f"   📅 Vencimento       : {item['vencimento']}")
            print(f"   ⏰ Status           : {item['status_vencimento']}")
            print("-"*60)

    if falhas:
        print(f"\n⚠️ {len(falhas)} imagem(ns) não puderam ser processadas:")
        for msg in falhas:
            print(f"   - {msg}")

    print("="*60)


# ==============================================================================
# EXECUÇÃO
# ==============================================================================

if __name__ == "__main__":
    pesquisar_imagem()
